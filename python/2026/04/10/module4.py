import openpyxl
from random import *

wb = openpyxl.load_workbook("example.xlsx")

ws1 = wb.create_sheet("lotto")

for y in range(1, 11):
  for x in range(1, 31):
    ws1.cell(x, y, randint(1, 100))

wb.save("example.xlsx")